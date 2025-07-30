import pandas as pd
from datetime import datetime
import os
import uuid
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

log_file = "trade_logs.xlsx"

def autosize_columns(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(filepath)

def log_trade_decision(price, granularity_list, trade_direction, tp, sl):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    trade_id = str(uuid.uuid4())
    data = {
        "trade_id": [trade_id],
        "timestamp": [timestamp],
        "price": [price],
        "granularities": [", ".join(granularity_list)],
        "trade_direction": [trade_direction],
        "take_profit": [tp],
        "stop_loss": [sl],
        "status": ["pending"],
        "exit_reason": [""],
        "balance": [""]
    }
    df_new = pd.DataFrame(data)

    try:
        if not os.path.exists(log_file):
            df_new.to_excel(log_file, index=False)
        else:
            df_existing = pd.read_excel(log_file)
            with pd.ExcelWriter(log_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                startrow = df_existing.shape[0] + 1
                df_new.to_excel(writer, index=False, header=False, startrow=startrow)
        autosize_columns(log_file)
        print(f"[Logger] 기록 저장 완료 → {log_file}")
        return trade_id
    except Exception as e:
        print(f"[Logger Error] 기록 실패: {e}")
        return None


def update_trade_exit(trade_id, new_status, exit_reason=None, exit_balance=None):
    if not os.path.exists(log_file):
        print("[Logger Error] 로그 파일이 없습니다.")
        return

    try:
        df = pd.read_excel(log_file)
        df.columns = [col.strip().lower() for col in df.columns]

        for col in ["exit_reason", "balance"]:
            if col not in df.columns:
                df[col] = ""

        updated = False
        for i in range(len(df)):
            if str(df.loc[i, "trade_id"]).strip() == trade_id.strip():
                df.at[i, "status"] = new_status
                if exit_reason:
                    df.at[i, "exit_reason"] = exit_reason
                if exit_balance:
                    df.at[i, "balance"] = exit_balance
                updated = True
                break

        if updated:
            df.to_excel(log_file, index=False)
            autosize_columns(log_file)
            print(f"[Logger] 업데이트 완료: {new_status}, 원인: {exit_reason}, 잔고: {exit_balance}")
        else:
            print(f"[Logger Warning] trade_id={trade_id} 에 해당하는 항목을 찾지 못했습니다.")
    except Exception as e:
        print(f"[Logger Error] 업데이트 중 예외 발생: {e}")
